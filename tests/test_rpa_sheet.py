import shutil
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

from app.utils.rpa_sheet import (
    RPA_HEADERS,
    build_rpa_row,
    parse_hive_table_comment,
    parse_location_uri,
    strip_hive_location_line,
    write_rpa_sheet,
)


class TestRpaSheetParsers(unittest.TestCase):
    def test_strip_hive_location_line_removes_location_row(self):
        sql = (
            "CREATE TABLE t (\n  `id` int\n)\n"
            "STORED AS RCFILE\n"
            "LOCATION 'viewfs://c9/dw/ods/foo';"
        )
        out = strip_hive_location_line(sql)
        self.assertNotIn("LOCATION", out)
        self.assertIn("STORED AS RCFILE", out)
        self.assertIn("CREATE TABLE", out)

    def test_strip_hive_location_line_case_insensitive(self):
        sql = "X\nlocation 'viewfs://a/b'\nY"
        out = strip_hive_location_line(sql)
        self.assertEqual(out, "X\nY")

    def test_parse_location_uri(self):
        sql = "PARTITIONED BY\nLOCATION 'viewfs://c9/dw/sds/tbl';"
        self.assertEqual(parse_location_uri(sql), "viewfs://c9/dw/sds/tbl")

    def test_parse_location_uri_missing(self):
        self.assertEqual(parse_location_uri("CREATE TABLE t (id int);"), "")

    def test_parse_hive_table_comment_simple(self):
        sql = (
            "CREATE EXTERNAL TABLE `default`.`x` (\n"
            "  `id` int COMMENT '主键'\n"
            ")\n"
            "COMMENT '测试表天表'\n"
            "PARTITIONED BY (`dt` string)\n"
        )
        self.assertEqual(parse_hive_table_comment(sql), "测试表天表")

    def test_parse_hive_table_comment_escaped_quote(self):
        sql = (
            "CREATE TABLE `default`.`x` (\n  `id` int\n)\n"
            "COMMENT 'It\\'s ok'\n"
            "PARTITIONED BY\n"
        )
        self.assertEqual(parse_hive_table_comment(sql), "It's ok")

    def test_parse_hive_table_comment_none(self):
        sql = "CREATE TABLE t (\n  `id` int\n)\nPARTITIONED BY\nLOCATION 'x';"
        self.assertEqual(parse_hive_table_comment(sql), "")


class TestBuildRpaRow(unittest.TestCase):
    def test_hive_row(self):
        sql = (
            "CREATE TABLE x (\n  `id` int\n)\n"
            "COMMENT 'c'\n"
            "STORED AS RCFILE\n"
            "LOCATION 'viewfs://c9/dw/ods/name';"
        )
        row = build_rpa_row(target_table_type="hive", create_sql=sql, dw_layer="ods")
        self.assertEqual(row["数据描述信息"], "c")
        self.assertEqual(row["数仓分层"], "ods")
        self.assertEqual(row["存储路径值"], "viewfs://c9/dw/ods/name")
        self.assertEqual(row["表类型"], "hive")
        self.assertNotIn("LOCATION", row["建表语句"])

    def test_clickhouse_row(self):
        sql = "CREATE LOCAL; CREATE DIST;"
        row = build_rpa_row(target_table_type="clickhouse", create_sql=sql, dw_layer="ignored")
        self.assertEqual(row["建表语句"], sql)
        self.assertEqual(row["数据描述信息"], "")
        self.assertEqual(row["数仓分层"], "")
        self.assertEqual(row["存储路径值"], "")
        self.assertEqual(row["表类型"], "clickhouse")


class TestWriteRpaSheet(unittest.TestCase):
    def setUp(self):
        self.tmp_dir = Path(tempfile.mkdtemp())

    def tearDown(self):
        shutil.rmtree(self.tmp_dir, ignore_errors=True)

    def test_write_preserves_tables_fields_and_replaces_rpa(self):
        path = self.tmp_dir / "book.xlsx"
        wb = Workbook()
        wb.active.title = "tables"
        wb.create_sheet("fields")
        wb.save(path)

        row = {
            "数据描述信息": "d",
            "数仓分层": "ods",
            "建表语句": "sql",
            "存储路径值": "viewfs://x",
            "表类型": "hive",
        }
        write_rpa_sheet(path, [row])
        wb2 = load_workbook(path)
        self.assertEqual(wb2.sheetnames, ["tables", "fields", "rpa"])
        ws = wb2["rpa"]
        self.assertEqual([c.value for c in ws[1]], RPA_HEADERS)
        self.assertEqual([c.value for c in ws[2]], ["d", "ods", "sql", "viewfs://x", "hive"])

        write_rpa_sheet(path, [])
        wb3 = load_workbook(path)
        self.assertEqual(wb3["rpa"].max_row, 1)

    def test_write_rpa_sheet_missing_file(self):
        with self.assertRaises(FileNotFoundError):
            write_rpa_sheet(self.tmp_dir / "nope.xlsx", [])

