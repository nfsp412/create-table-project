"""RPA 汇总 sheet 写入 create_table_info.xlsx。"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Any, Mapping

from openpyxl import Workbook, load_workbook

RPA_HEADERS = ["数据描述信息", "数仓分层", "建表语句", "存储路径值", "表类型"]


def strip_hive_location_line(sql: str) -> str:
    """去掉 Hive DDL 中含 LOCATION 的整行（与生成器中单行 LOCATION 一致）。"""
    lines = sql.splitlines()
    out: list[str] = []
    for line in lines:
        stripped = line.strip()
        if re.match(r"(?i)^LOCATION\s+'", stripped):
            continue
        out.append(line)
    return "\n".join(out)


def parse_location_uri(sql: str) -> str:
    """从 DDL 中提取第一段 LOCATION 后的单引号内路径。"""
    m = re.search(r"(?i)LOCATION\s+'([^']+)'", sql)
    return m.group(1).strip() if m else ""


def parse_hive_table_comment(sql: str) -> str:
    """
    从 Hive CREATE TABLE DDL 中解析表级 COMMENT（列定义之后的 COMMENT 子句）。
    支持注释内的 \\' 转义。
    """
    m = re.search(
        r"\)\s*\n\s*COMMENT\s+'((?:\\'|[^'])*)'",
        sql,
        re.IGNORECASE | re.DOTALL,
    )
    if not m:
        return ""
    return m.group(1).replace("\\'", "'")


def build_rpa_row(
    *,
    target_table_type: str,
    create_sql: str,
    dw_layer: str,
) -> dict[str, str]:
    """生成一条 rpa 行（仅用于新建表分支）。"""
    if target_table_type == "clickhouse":
        return {
            "数据描述信息": "",
            "数仓分层": "",
            "建表语句": create_sql,
            "存储路径值": "",
            "表类型": target_table_type,
        }
    return {
        "数据描述信息": parse_hive_table_comment(create_sql),
        "数仓分层": dw_layer,
        "建表语句": strip_hive_location_line(create_sql),
        "存储路径值": parse_location_uri(create_sql),
        "表类型": target_table_type,
    }


def write_rpa_sheet(excel_path: Path, rows: list[Mapping[str, Any]]) -> None:
    """
    在 xlsx 上新建或覆盖 `rpa` sheet；文件不存在时创建新工作簿（仅含 rpa sheet）。
    已存在文件时保留除 `rpa` 外的其他 sheet。
    """
    path = Path(excel_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    if path.is_file():
        wb = load_workbook(path)
    else:
        wb = Workbook()
        wb.remove(wb.active)
    if "rpa" in wb.sheetnames:
        wb.remove(wb["rpa"])
    ws = wb.create_sheet("rpa")
    ws.append(RPA_HEADERS)
    for row in rows:
        ws.append([str(row.get(h, "") or "") for h in RPA_HEADERS])
    wb.save(path)
